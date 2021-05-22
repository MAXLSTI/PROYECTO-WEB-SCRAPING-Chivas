from openpyxl import Workbook
import os
#En esta parte mandamos a llamar los nombres de losjugadores en el archivo correspondiente
def excel():
    archivo=open("nombres jugadores.txt","r")
    lineas=archivo.readlines()
    data=[]
    for i in lineas:
        data.append(i[:-1])
    nombres=[]
    for linea in data:
        if linea == "José Juan Macías" or linea=="Uriel Antuna" or linea=="Isaac Brizuela" or linea=="Jesus Angulo" or linea=="Jesus Sánchez":
            nombres.append(linea)
    archivo.close()
##Lectura de datos por archivo para vaciar al excel de los 4 archivos dr la historia
    archivo=open("El origen.txt","r")
    lineas=archivo.readlines()
    info=[]
    for i in lineas:
        info.append(i)
    archivo.close()
    lec=len(info)
##
    archivo=open("Vergara y el renacer.txt","r")
    lineas=archivo.readlines()
    infoVYR=[]
    for i in lineas:
        infoVYR.append(i)
    archivo.close()
    lecVYR=len(infoVYR)
##
    archivo=open("La Epoca Reciente.txt","r")
    lineas=archivo.readlines()
    infoLI=[]
    for i in lineas:
        infoLI.append(i)
    archivo.close()
    lecLI=len(infoLI)
##
    archivo=open("Campeonismo.txt","r")
    lineas=archivo.readlines()
    infoCAP=[]
    for i in lineas:
        infoCAP.append(i)
    archivo.close()
    lecCAP=len(infoCAP)
##
    archivo=open("José Juan Macías.txt","r")
    lineas=archivo.readlines()
    infoJJM=[]
    for i in lineas:
        infoJJM.append(i)
    archivo.close()
    lecJJM=len(infoJJM)
##
    archivo=open("Uriel Antuna.txt","r")
    lineas=archivo.readlines()
    infoUA=[]
    for i in lineas:
        infoUA.append(i)
    archivo.close()
    lecUA=len(infoUA)
##
    archivo=open("Isaac Brizuela.txt","r")
    lineas=archivo.readlines()
    infoIB=[]
    for i in lineas:
        infoIB.append(i)
    archivo.close()
    lecIB=len(infoIB)
##
    archivo=open("Jesus Angulo.txt","r")
    lineas=archivo.readlines()
    infoJA=[]
    for i in lineas:
        infoJA.append(i)
    archivo.close()
    lecJA=len(infoJA)
##
    archivo=open("Jesus Sánchez.txt","r")
    lineas=archivo.readlines()
    infoJS=[]
    for i in lineas:
        infoJS.append(i)
    archivo.close()
    lecJS=len(infoJS)

    ##
    archivo=open("Fechas partidos.txt","r")
    lineas=archivo.readlines()
    infoFP=[]
    for i in lineas:
        infoFP.append(i)
    archivo.close()
    lecFP=len(infoFP)
    ##
    archivo=open("Hora partidos.txt","r")
    lineas=archivo.readlines()
    infoHP=[]
    for i in lineas:
        infoHP.append(i)
    archivo.close()
    lecHP=len(infoHP)
    ##
    archivo=open("Marcador final.txt","r")
    lineas=archivo.readlines()
    infoMF=[]
    for i in lineas:
        infoMF.append(i)
    archivo.close()
    lecMF=len(infoMF)
    ##
    archivo=open("Liga.txt","r")
    lineas=archivo.readlines()
    infoLA=[]
    for i in lineas:
        infoLA.append(i)
    archivo.close()
    lecLA=len(infoLA)
#inicia el archivo excel
    libro = Workbook()#work(): crea un libro y lo almacenamos en "libro", que será nuestro libro

    pagina= libro.active



    pagina1 = libro.create_sheet("José Juan Macías")
    pagina2=libro.create_sheet("Uriel Antuna")
    pagina3 = libro.create_sheet("Isaac Brizuela")
    pagina4=libro.create_sheet("Jesus Angulo")
    pagina5=libro.create_sheet("Jesus Sánchez")
    pagina.title="El origen de chivas"
    pagina6=libro.create_sheet("Campeonismo")
    pagina7=libro.create_sheet("La Epoca Reciente")
    pagina8=libro.create_sheet("Vergara y el renacer")
    pagina9=libro.create_sheet("Tabla de Partidos")
## Se realiza el llenado del excel en cada una de sus paginas
    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lec):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina[i]=info[q]
        q+=1
##
    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecJJM):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina1[i]=infoJJM[q]
        q+=1
##
    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecUA):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina2[i]=infoUA[q]
        q+=1
##
    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecIB):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina3[i]=infoIB[q]
        q+=1
##
    m="A"
    c=1
    suma=m+str(c)
#c=1

    list=[]
    for i in range(lecJA):
        su=m+str(c)
        list.append(su)
        c+=1
    q=-2
    for i in list:
        pagina4[i]=infoJA[q]
        q+=1
##

    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecJS):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina5[i]=infoJS[q]
        q+=1


##
    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecCAP):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina6[i]=infoCAP[q]
        q+=1

##
    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecLI):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina7[i]=infoLI[q]
        q+=1

##
    m="A"
    c=1
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecVYR):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina8[i]=infoVYR[q]
        q+=1
##
    m="A"
    c=2
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecFP):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina9[i]=infoFP[q]
        q+=1


##
    m="C"
    c=2
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecHP):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina9[i]=infoHP[q]
        q+=1
##
    m="D"
    c=2
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecMF):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina9[i]=infoMF[q]
        q+=1
##
    m="G"
    c=2
    suma=m+str(c)
#c=1
    list=[]
    for i in range(lecLA):
        su=m+str(c)
        list.append(su)
        c+=1
    q=0
    for i in list:
        pagina9[i]=infoLA[q]
        q+=1
##
##


## LLenado de los nombres de los jugadores de  cada pagina
    pagina1['A1'] = nombres[0]
    pagina2['A1'] = nombres[1]
    pagina3['A1'] = nombres[2]
    pagina4['A1'] = nombres[3]
    pagina5['A1'] = nombres[4]
    pagina9['B1'] = "TABLA DE ULTIMOS PARTIDOS 2021"
######################################################
#print(libro.sheetnames)
    libro.save("Bob Esponja.xlsx")
